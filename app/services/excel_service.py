"""
Excel processing service for data enrichment and file manipulation
"""

import pandas as pd
import io
import tempfile
import os
import logging
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..core.config import SAMPLE_RULES
from ..models.schemas import VehicleInfo, EnrichmentResult
from ..utils.excel_utils import find_column_mapping, extract_vehicle_info
from ..utils.formatting_utils import (
    copy_cell_style, apply_cell_style, auto_adjust_column_widths,
    get_data_row_styles, clear_data_rows, find_original_table_end
)
from .openai_service import OpenAIService

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for Excel processing operations"""
    
    @staticmethod
    def apply_ai_enrichment(df: pd.DataFrame, rules: Dict[str, Any] = None) -> pd.DataFrame:
        """
        Apply AI-based enrichment rules to the DataFrame using OpenAI for intelligent classification
        """
        if rules is None:
            rules = SAMPLE_RULES
            
        # Create a copy to avoid modifying the original
        enriched_df = df.copy()
        
        # Get the reference column for mapping with fuzzy matching
        target_reference_column = rules["reglas_asignacion"]["mapeo_columnas"]["columna_referencia"]
        
        try:
            # Find the actual column name using fuzzy matching
            reference_column = find_column_mapping(enriched_df, target_reference_column)
            logger.info(f"Using column '{reference_column}' for vehicle type classification")
        except ValueError as e:
            logger.error(f"Column mapping error: {str(e)}")
            raise ValueError(f"Reference column '{target_reference_column}' not found in data. Available columns: {list(enriched_df.columns)}")
        
        # Add new columns for enrichment
        new_columns = rules["reglas_asignacion"]["columnas_a_agregar"]
        for col in new_columns:
            enriched_df[col] = ""
        
        # Get available vehicle types from rules
        available_types = list(rules["coberturas_por_tipo"].keys())
        
        # Apply enrichment rules based on AI-powered vehicle classification
        for index, row in enriched_df.iterrows():
            vehicle_description = str(row[reference_column]).strip()
            
            if not vehicle_description or vehicle_description.lower() in ['nan', 'none', '']:
                logger.warning(f"Empty vehicle description at row {index}, skipping")
                continue
            
            try:
                # Use OpenAI to classify the vehicle type
                coverage_type = OpenAIService.classify_vehicle(vehicle_description, available_types)
                logger.info(f"Row {index}: '{vehicle_description}' classified as '{coverage_type}'")
                
                # Extract additional vehicle information
                vehicle_extra_info = extract_vehicle_info(row, enriched_df)
                
                # Create vehicle info object
                vehicle_info = VehicleInfo(
                    description=vehicle_description,
                    type=coverage_type,
                    year=vehicle_extra_info["year"],
                    model=vehicle_extra_info["model"]
                )
                
                # Get coverage rules for this vehicle type
                if coverage_type in rules["coberturas_por_tipo"]:
                    coverages = rules["coberturas_por_tipo"][coverage_type]["coberturas"]
                    
                    # Generate specific values for DANOS MATERIALES
                    if "DANOS MATERIALES" in coverages:
                        danos_values = OpenAIService.generate_insurance_values(
                            vehicle_info, "DANOS MATERIALES"
                        )
                        enriched_df.at[index, "DANOS MATERIALES LIMITES"] = danos_values.LIMITES
                        enriched_df.at[index, "DANOS MATERIALES DEDUCIBLES"] = danos_values.DEDUCIBLES
                        logger.info(f"Row {index} DANOS MATERIALES: Limites={danos_values.LIMITES}, Deducibles={danos_values.DEDUCIBLES}")
                    
                    # Generate specific values for ROBO TOTAL
                    if "ROBO TOTAL" in coverages:
                        robo_values = OpenAIService.generate_insurance_values(
                            vehicle_info, "ROBO TOTAL"
                        )
                        enriched_df.at[index, "ROBO TOTAL LIMITES"] = robo_values.LIMITES
                        enriched_df.at[index, "ROBO TOTAL DEDUCIBLES"] = robo_values.DEDUCIBLES
                        logger.info(f"Row {index} ROBO TOTAL: Limites={robo_values.LIMITES}, Deducibles={robo_values.DEDUCIBLES}")
                else:
                    logger.warning(f"Coverage type '{coverage_type}' not found in rules")
                    
            except Exception as e:
                logger.error(f"Error processing row {index}: {str(e)}")
                # Continue with next row instead of failing completely
                continue
        
        return enriched_df
    
    @staticmethod
    def create_enriched_excel(
        original_content: bytes, 
        enriched_df: pd.DataFrame, 
        sheet_name: str, 
        header_row: int,
        rules: Dict[str, Any] = None
    ) -> str:
        """
        Create an enriched Excel file with proper formatting
        """
        if rules is None:
            rules = SAMPLE_RULES
            
        # Create a temporary file for the modified Excel
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_filename = temp_file.name
        temp_file.close()
        
        try:
            # Load the original workbook to preserve all formatting
            wb = load_workbook(io.BytesIO(original_content))
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Find where data starts (after header row)
                data_start_row = header_row + 2  # +1 for 0-based to 1-based, +1 for row after header
                header_row_excel = header_row + 1  # Convert to 1-based for Excel
                
                # Find the actual end of the original table data
                original_table_end_col = find_original_table_end(ws, header_row_excel)
                
                # Map existing columns
                col_mapping = {}
                for col_idx in range(1, original_table_end_col + 1):
                    header_cell = ws.cell(row=header_row_excel, column=col_idx)
                    if header_cell.value:
                        col_name = str(header_cell.value).strip()
                        col_mapping[col_name] = col_idx
                
                # Add new columns right after the original table
                new_columns = rules["reglas_asignacion"]["columnas_a_agregar"]
                next_col = original_table_end_col + 1
                
                # Get header style from the last original column to copy formatting
                last_original_header = ws.cell(row=header_row_excel, column=original_table_end_col)
                header_style = copy_cell_style(last_original_header)
                
                for new_col in new_columns:
                    if new_col not in col_mapping:
                        # Add the new column header with same formatting as original headers
                        new_header_cell = ws.cell(row=header_row_excel, column=next_col, value=new_col)
                        apply_cell_style(new_header_cell, header_style)
                        
                        col_mapping[new_col] = next_col
                        logger.info(f"Added new column '{new_col}' at position {next_col} ({get_column_letter(next_col)})")
                        next_col += 1
                
                # Get data row styles for copying formatting
                data_row_styles = get_data_row_styles(ws, data_start_row, original_table_end_col)
                
                # Clear only the data rows, not headers
                max_row = ws.max_row
                clear_data_rows(ws, data_start_row, max_row, next_col)
                
                # Write the enriched data while preserving formatting
                for row_idx, (df_idx, df_row) in enumerate(enriched_df.iterrows()):
                    excel_row = data_start_row + row_idx
                    
                    # Write data for each column
                    for col_name, col_pos in col_mapping.items():
                        if col_name in enriched_df.columns:
                            value = df_row[col_name]
                            # Handle NaN values
                            if pd.isna(value):
                                value = ""
                            
                            cell = ws.cell(row=excel_row, column=col_pos, value=value)
                            
                            # Apply formatting based on whether it's an original or new column
                            if col_pos <= original_table_end_col and col_pos in data_row_styles:
                                # Use original formatting for existing columns
                                apply_cell_style(cell, data_row_styles[col_pos])
                            elif col_pos > original_table_end_col:
                                # Use similar formatting for new columns (based on last original column)
                                if original_table_end_col in data_row_styles:
                                    apply_cell_style(cell, data_row_styles[original_table_end_col])
                
                # Auto-adjust column widths for new columns
                auto_adjust_column_widths(ws, original_table_end_col + 1, next_col)
                
                logger.info(f"Updated sheet '{sheet_name}' with {len(enriched_df)} rows and {len(col_mapping)} columns")
                logger.info(f"Original table ended at column {original_table_end_col}, new columns start at {original_table_end_col + 1}")
                logger.info(f"Column mapping: {col_mapping}")
            
            # Save the modified workbook
            wb.save(temp_filename)
            wb.close()
            
            return temp_filename
            
        except Exception as e:
            # Clean up temp file if there's an error
            if os.path.exists(temp_filename):
                try:
                    os.unlink(temp_filename)
                except Exception as cleanup_error:
                    logger.error(f"Error cleaning up temp file: {cleanup_error}")
            raise e
