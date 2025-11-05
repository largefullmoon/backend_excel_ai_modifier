"""
Excel formatting utilities
"""

import logging
import copy
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from ..core.config import settings

logger = logging.getLogger(__name__)


def copy_cell_style(source_cell):
    """
    Copy the style from a source cell
    """
    return {
        'font': copy.copy(source_cell.font) if source_cell.font else Font(),
        'fill': copy.copy(source_cell.fill) if source_cell.fill else PatternFill(),
        'border': copy.copy(source_cell.border) if source_cell.border else Border(),
        'alignment': copy.copy(source_cell.alignment) if source_cell.alignment else Alignment()
    }


def apply_cell_style(cell, style_dict):
    """
    Apply a style dictionary to a cell
    """
    cell.font = style_dict['font']
    cell.fill = style_dict['fill']
    cell.border = style_dict['border']
    cell.alignment = style_dict['alignment']


def auto_adjust_column_widths(worksheet, start_col: int, end_col: int, width: int = None):
    """
    Auto-adjust column widths for specified columns
    """
    if width is None:
        width = settings.NEW_COLUMN_WIDTH
        
    for col_idx in range(start_col, end_col):
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = width
        logger.debug(f"Set column {col_letter} width to {width}")


def get_data_row_styles(worksheet, data_start_row: int, original_table_end_col: int):
    """
    Extract styles from the first data row for copying to new rows
    """
    data_row_styles = {}
    if worksheet.max_row >= data_start_row:
        for col_idx in range(1, original_table_end_col + 1):
            sample_cell = worksheet.cell(row=data_start_row, column=col_idx)
            data_row_styles[col_idx] = copy_cell_style(sample_cell)
    return data_row_styles


def clear_data_rows(worksheet, data_start_row: int, max_row: int, max_col: int):
    """
    Clear data rows while preserving headers
    """
    for row in range(data_start_row, max_row + 1):
        for col in range(1, max_col):
            worksheet.cell(row=row, column=col).value = None
    logger.info(f"Cleared data rows from {data_start_row} to {max_row}")


def find_original_table_end(worksheet, header_row_excel: int):
    """
    Find the last column with data in the original table
    """
    original_table_end_col = 1
    for col_idx in range(1, worksheet.max_column + 1):
        header_cell = worksheet.cell(row=header_row_excel, column=col_idx)
        if header_cell.value and str(header_cell.value).strip():
            original_table_end_col = col_idx
    
    logger.info(f"Original table ends at column {original_table_end_col} ({get_column_letter(original_table_end_col)})")
    return original_table_end_col
